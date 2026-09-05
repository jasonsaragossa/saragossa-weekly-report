"""
Import finance's monthly commission spreadsheets into the manual ledgers.

Two workbooks, one shape each:

  "Contract Commission - <Mon YY>.xlsx"      -> the Contract Entry ledger
      sheet "Commission Report": one row per contractor per week, with a
      Contribution column. Territory sheets are slices of the same rows, so
      only the master sheet is read.

  "Deploy & Component Summary - <Mon YY>.xlsx" -> the Deploy & Component ledger
      sheets "Deploy & Component" and "Deploy & Component US": blocks per
      consultant with the header row REPEATED before each block, so the column
      positions have to be re-read as we go.

Figures are taken in their native currency (US desks in USD), matching how the
ledgers already store them. Names are matched to Mercury on a normalised
full name, including disabled users so leavers still get their contribution.
"""
import io
import re
from collections import defaultdict
from datetime import date

CONTRACT_SHEET = "Commission Report"
DEPLOY_SHEETS = ("Deploy & Component", "Deploy & Component US")
NO_CONSULTANT = "(no consultant on the row)"
MONTHS = ("jan", "feb", "mar", "apr", "may", "jun",
          "jul", "aug", "sep", "oct", "nov", "dec")
CONTRIBUTION = "contribution"
CONSULTANT = "consultant"
# Rows that are structural rather than data
SKIP_VALUES = {"consultant", "january commission", "february commission",
               "march commission", "april commission", "may commission",
               "june commission", "july commission", "august commission",
               "september commission", "october commission",
               "november commission", "december commission", "total"}


def month_from_filename(filename: str):
    """
    (year, month) from "Contract Commission - Jan 26.xlsx", or None.

    The month is taken from the name rather than the rows: 'Period End' on the
    contract sheet is the commission PERIOD and legitimately spans four months
    either side of the run, in three different date formats. The workbook as a
    whole is one month's commission run, which is what the filename states.
    """
    text = filename.lower()
    for i, mon in enumerate(MONTHS, 1):
        m = re.search(rf"{mon}[a-z]*[\s\-_]*(\d{{2,4}})", text)
        if m:
            year = int(m.group(1))
            if year < 100:
                year += 2000
            if 2000 <= year <= date.today().year + 1:
                return year, i
    return None


def normalise_name(value) -> str:
    """Letters only, lowercased — survives 'Liam Mustoe- Linnane' vs 'Liam Mustoe-Linnane'."""
    return re.sub(r"[^a-z]", "", str(value or "").lower())


def _header_map(row) -> dict:
    """{lowercased header: index} for a header row, or {} if it isn't one."""
    cells = [str(c).strip().lower() if c is not None else "" for c in row]
    if CONSULTANT in cells and CONTRIBUTION in cells:
        # 'Consultant' appears twice on the Deploy sheets (owner and contractor);
        # the first is the owner, which is the one we want.
        return {name: i for i, name in reversed(list(enumerate(cells))) if name}
    return {}


def _rows_from_sheet(ws, allow_unnamed: bool) -> list:
    """(name, contribution) pairs, re-reading the header wherever it repeats."""
    out, cols = [], {}
    for row in ws.iter_rows(values_only=True):
        header = _header_map(row)
        if header:
            cols = header
            continue
        if not cols:
            continue
        name = row[cols[CONSULTANT]] if cols.get(CONSULTANT) is not None else None
        name = str(name).strip() if name is not None else ""
        if name.lower() in SKIP_VALUES:
            continue
        raw = row[cols[CONTRIBUTION]] if cols.get(CONTRIBUTION) is not None else None
        if not isinstance(raw, (int, float)) or isinstance(raw, bool):
            continue
        if not name:
            # On the contract sheet a blank owner is a real row whose owner cell
            # was left empty, so it is kept under a visible label rather than
            # silently dropped. On the deploy sheets a blank owner means the
            # "Commission Payable" subtotal that closes each consultant's block
            # — counting those would double the workbook.
            if not allow_unnamed:
                continue
            name = NO_CONSULTANT
        out.append((name, float(raw)))
    return out


def parse_workbook(data: bytes) -> dict:
    """
    {"kind": "contract"|"solution", "totals": {name: contribution},
     "rows": n, "sheets_used": [...], "sheets_skipped": [...]}
    Raises ValueError if the workbook isn't one of the two known shapes.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    names = set(wb.sheetnames)

    if CONTRACT_SHEET in names:
        # The territory sheets are slices of the master, so reading only the
        # master avoids double-counting.
        kind, sheets = "contract", [CONTRACT_SHEET]
    elif names & set(DEPLOY_SHEETS):
        kind = "solution"
        sheets = [s for s in DEPLOY_SHEETS if s in names]
    else:
        raise ValueError(
            "Unrecognised workbook — expected a 'Commission Report' sheet (contract) "
            f"or a 'Deploy & Component' sheet. Found: {', '.join(sorted(names))}")

    totals, count = defaultdict(float), 0
    for sheet in sheets:
        for name, value in _rows_from_sheet(wb[sheet], allow_unnamed=(kind == "contract")):
            totals[name] += value
            count += 1
    wb.close()
    if not totals:
        raise ValueError("No Contribution figures found in that workbook.")
    return {"kind": kind, "totals": {k: round(v, 2) for k, v in totals.items()},
            "rows": count, "sheets_used": sheets,
            "sheets_skipped": sorted(names - set(sheets))}


def match_to_users(totals: dict, users: list) -> dict:
    """
    {"matched": [{uid, name, sheet_name, amount, disabled}], "unmatched": [{name, amount}]}
    Disabled users still match — leavers keep their contribution.
    """
    index = {}
    for u in users:
        key = normalise_name(u.get("fullname"))
        # Prefer an active user where two accounts share a name
        if key and (key not in index or (index[key].get("isdisabled") and not u.get("isdisabled"))):
            index[key] = u
    matched, unmatched = [], []
    for name, amount in sorted(totals.items(), key=lambda kv: -kv[1]):
        user = index.get(normalise_name(name))
        if user:
            matched.append({"uid": user["systemuserid"], "name": user.get("fullname"),
                            "sheet_name": name, "amount": amount,
                            "disabled": bool(user.get("isdisabled"))})
        else:
            unmatched.append({"name": name, "amount": amount})
    return {"matched": matched, "unmatched": unmatched}
